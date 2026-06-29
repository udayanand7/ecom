"""
Data-driven helpers to read test data from Excel (.xlsx) and CSV files.
"""

import csv
import os
from typing import Any, Dict, List

import openpyxl

from utils.logger import get_logger

log = get_logger("data_reader")


# ── Excel ─────────────────────────────────────────────────────────────────────

def read_excel(filepath: str, sheet_name: str = None) -> List[Dict[str, Any]]:
    """
    Read an Excel sheet and return a list of dicts (one per data row).
    The first row is treated as headers.
    """
    if not os.path.exists(filepath):
        log.error(f"Excel file not found: {filepath}")
        return []

    wb    = openpyxl.load_workbook(filepath, data_only=True)
    ws    = wb[sheet_name] if sheet_name else wb.active
    rows  = list(ws.iter_rows(values_only=True))

    if not rows:
        return []

    headers = [str(h).strip() if h is not None else f"col_{i}"
               for i, h in enumerate(rows[0])]

    data = []
    for row in rows[1:]:
        if all(cell is None for cell in row):
            continue                          # skip blank rows
        record = {headers[i]: row[i] for i in range(len(headers))}
        data.append(record)

    log.info(f"Read {len(data)} rows from '{filepath}' "
             f"(sheet: {ws.title})")
    return data


def write_excel(filepath: str, data: List[Dict[str, Any]],
                sheet_name: str = "Results") -> None:
    """Write a list of dicts back to an Excel file."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name

    if not data:
        wb.save(filepath)
        return

    headers = list(data[0].keys())
    ws.append(headers)
    for record in data:
        ws.append([record.get(h) for h in headers])

    wb.save(filepath)
    log.info(f"Wrote {len(data)} rows → {filepath}")


# ── CSV ───────────────────────────────────────────────────────────────────────

def read_csv(filepath: str) -> List[Dict[str, Any]]:
    """Read a CSV file and return a list of dicts."""
    if not os.path.exists(filepath):
        log.error(f"CSV file not found: {filepath}")
        return []

    with open(filepath, newline="", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        data   = [dict(row) for row in reader]

    log.info(f"Read {len(data)} rows from '{filepath}'")
    return data
